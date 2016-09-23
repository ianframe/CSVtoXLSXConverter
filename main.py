import csv
import os
from openpyxl import Workbook

################################################

# convert function
def convertCSVtoXLSX(fileName, pathDir):
    print("Current CSV file: " + fileName)
    wb.create_sheet(fileName)
    ws = wb.get_sheet_by_name(fileName)
    ws.title = fileName

    quakeDataCSV = open(pathDir + '/' + fileName) # returns a CSV file object
    csvReader = csv.reader(quakeDataCSV)  # create the Reader object to iterate over the rows
    csvData = list(csvReader)

    rowIndex = 1  # used to indicate the row to fill a value in our XLS file
    colIndex = ['A', 'B', 'C']  # used to index the appropriate column  in our XLS file

    for row in csvData:
        row = row[0:3]
        colNum = 0  # reset the csv colIndex file for each new row
        for val in row:
            ws[colIndex[colNum] + str(rowIndex)] = val
            colNum += 1
        rowIndex += 1

    wb.save('QuakeData.xlsx')

################################################

# establish the directory path
path = './CSVfiles'
files = os.listdir(path)

wb = Workbook(guess_types=True)  # create a new workbook
wb.remove_sheet(wb.active)

for file in files:
    convertCSVtoXLSX(file, path)
    # print(file)

print("File compiltation and conversion is complete.")